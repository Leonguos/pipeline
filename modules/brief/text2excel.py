#!/usr/bin/env python
# -*- coding=utf-8 -*-
import re
import os
import sys
import argparse
import codecs
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.dirname(os.path.dirname(BASE_DIR)))

from utils import mkdir_if_not_exists, safe_open


def get_sheetname(filename):
    
    filename = os.path.basename(filename)

    file_list = (
        'stat', 'readme',
        'roh_anno_snp', 'roh_anno',                                          # roh
        'AllGene_list', 'CandidateGene_list', 'CandidateGene_score',         # phenolyzer
        'Gene_interactions', 'Gene_description', 'Networks.description',     # ppi
        'go_CC', 'go_BP', 'go_MF', 'kegg',                                   # pathway
        'intersect', 'denovoSV', 'denovoCNV',                                # denovo
        'freq.func.syn.deleterious',                                         # filterdb
        'Pathogenic', 'LikelyPathogenic', 'VUS', 'LikelyBenign', 'Benign',   # acmg
        'LikelyDeleterious',                                                 # filter_sv, filter_cnv
        'noncoding',                                                         # noncoding
        'CandidateGene', 'dominant', 'recessive',                            # model
        'PatientShare',                                                      # share
        'LinkageAnalysis',                                                   # linkage
    )

    sheetname = ''
    for each in file_list:
        if each in filename:
            sheetname = each
            break

    if not sheetname:
        sheetname = re.sub(r'(\.xls)|(\.gz)', '', filename).split('.')[-1]

    return sheetname[:31]


def text2excel(outfile, *infiles):
    '''
    infiles can be a string: 'a.xls,b.xls'          ('a.xls,b.xls', )
                  or a list: ['a.xls', 'b.xls']     (['a.xls', 'b.xls'], )
    or many positional args: 'a.xls', 'b.xls'       ('a.xls', 'b.xls')
    '''

    # wb = openpyxl.Workbook(encoding='utf8')
    wb = openpyxl.Workbook()

    if len(infiles) == 1:
        if isinstance(infiles[0], str):
            infile_list = infiles[0].split(',')
        elif isinstance(infiles[0], list):
            infile_list = infiles[0]

    elif len(infiles) >= 2:
        infile_list = list(infiles)

    else:
        exit('error infiles: {}'.format(infiles))

    for infile in infile_list:
        sheetname = get_sheetname(infile)
        print 'create sheet:', sheetname
        sheet = wb.create_sheet(title=sheetname)
        # with codecs.open(infile, mode='r', encoding='gbk', errors='ignore') as f:
        with safe_open(infile) as f:
            for n, line in enumerate(f):
                row = n + 1
                linelist = line.strip().split('\t')
                for m, value in enumerate(linelist):
                    column = m + 1
                    sheet.cell(row=row, column=column, value=value)

    # remove default sheet
    try:
        wb.remove(wb['Sheet'])
    except AttributeError:
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))  # for the old version

    outdir = os.path.dirname(outfile)
    if outdir:
        mkdir_if_not_exists(outdir)

    wb.save(filename=outfile)

    print 'write excel file:', outfile


if __name__ == "__main__":

    usage = '''usage:
    python %s <outfile> <infiles>
        outfile   the output excel format file
        infiles   the input text format files, separate by comma
    ''' % sys.argv[0]

    if len(sys.argv) < 3:
        print usage
        exit(1)

    text2excel(sys.argv[1], *sys.argv[2:])
